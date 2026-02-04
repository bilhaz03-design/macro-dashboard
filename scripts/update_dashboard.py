#!/usr/bin/env python3
import json
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path('/Users/bobbo/Desktop/Finans Projects')
DATA_JSON = ROOT / 'dashboard' / 'data.json'
XLSX_PATH = ROOT / 'output' / 'spreadsheet' / 'macro_dashboard.xlsx'
HTML_PATH = ROOT / 'dashboard' / 'index.html'


def load_data():
    return json.loads(DATA_JSON.read_text(encoding='utf-8'))


def update_excel(data):
    wb = openpyxl.load_workbook(XLSX_PATH)
    overview = wb['Overview']

    # Update overview date
    overview['A2'] = f"Last updated: {data.get('last_updated', date.today().isoformat())}"

    # Update overview country rows (assumes Sweden row 5, China row 6)
    country_map = {"Sweden": 5, "China": 6}
    for c in data.get('summary', []):
        row = country_map.get(c.get('title'))
        if not row:
            continue
        overview[f'B{row}'] = c.get('value', '')
        overview[f'C{row}'] = c.get('note', '')
        overview[f'D{row}'] = 'Weekly'

    # Global risk section
    if 'global_risk' in data:
        overview['G5'] = data['global_risk'].get('status', '')
        overview['G6'] = data['global_risk'].get('note', '')

    # Update country sheets
    for country in data.get('countries', []):
        name = country.get('name')
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        ws['A3'] = 'Last update'
        ws['B3'] = data.get('last_updated', date.today().isoformat())
        # Build label -> row map from column A
        label_rows = {ws[f'A{r}'].value: r for r in range(5, 13)}
        for block in country.get('blocks', []):
            row = label_rows.get(block.get('label'))
            if not row:
                continue
            ws[f'B{row}'] = block.get('status', '')
            ws[f'C{row}'] = block.get('detail', '')
            ws[f'D{row}'] = ''
            ws[f'E{row}'] = ''
            # first source URL if present
            sources = block.get('sources') or []
            if sources:
                ws[f'F{row}'] = sources[0].get('url', '')

    wb.save(XLSX_PATH)

def update_html_inline(data):
    if not HTML_PATH.exists():
        return
    html = HTML_PATH.read_text(encoding='utf-8')
    marker_start = '<script type="application/json" id="inline-data">'
    marker_end = '</script>'
    if marker_start not in html:
        return
    before, rest = html.split(marker_start, 1)
    _, after = rest.split(marker_end, 1)
    inline = json.dumps(data, indent=2)
    HTML_PATH.write_text(before + marker_start + inline + marker_end + after, encoding='utf-8')

def main():
    if not DATA_JSON.exists():
        raise SystemExit(f"Missing {DATA_JSON}")
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing {XLSX_PATH}")

    data = load_data()
    update_excel(data)
    update_html_inline(data)
    print("Updated Excel from dashboard/data.json")


if __name__ == '__main__':
    main()
